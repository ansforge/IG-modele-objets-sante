#!/usr/bin/env python3
"""
Generate MOS.xlsx by parsing FSH logical model source files directly.

Reads  : sections.json  (Partie → [class IDs])
         input/fsh/logicals/**/*.fsh
Outputs: input/images/MOS.xlsx  (copied verbatim to IG output by the publisher)
"""

import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
IG_ROOT = SCRIPT_DIR.parent
FSH_DIR = IG_ROOT / "input" / "fsh" / "logicals"
OUTPUT_FILE = IG_ROOT / "input" / "images" / "MOS.xlsx"
SECTIONS_FILE = SCRIPT_DIR / "sections.json"

HEADERS = [
    "Partie", "Classe", "Attribut", "Cardinalité", "Type",
    "Description", "Nomenclature", "FHIR", "SEMIC", "HL7",
    "XDS", "CDA R2", "DICOM", "IHE PAM Fr", "OMOP",
]

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
H_FONT = Font(bold=True, color="FFFFFF")
H_FILL = PatternFill("solid", fgColor="1F4E79")
P_FONT = Font(bold=True, color="FFFFFF")
P_FILL = PatternFill("solid", fgColor="2E75B6")
C_FONT = Font(bold=True)
C_FILL = PatternFill("solid", fgColor="BDD7EE")
ODD_FILL = PatternFill("solid", fgColor="EEF4FB")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")

# ---------------------------------------------------------------------------
# FSH parsing
# ---------------------------------------------------------------------------

_Q = r'"((?:[^"\\]|\\.)*)"'  # capture a quoted FSH string
_ELEM = re.compile(
    r'^\*\s+([A-Za-z][A-Za-z0-9_.]*)\s+(\d+\.\.[0-9*]+)\s+(\S+)\s+' + _Q + r'\s+' + _Q
)
_BIND = re.compile(r'^\*\s+([A-Za-z][A-Za-z0-9_.]*)\s+from\s+(\S+)')
_DESC = re.compile(r'^Description:\s+' + _Q, re.MULTILINE)


def _nos_name(vs_url: str) -> str:
    """Extract the NOS/TRE identifier from a ValueSet URL."""
    m = re.search(r"/NOS/([^/]+)/FHIR/", vs_url)
    return m.group(1) if m else vs_url.rstrip("?vs").split("/")[-1]


def _type_name(code: str) -> str:
    return code.split("/")[-1] if code.startswith("http") else code


def parse_fsh(path: Path):
    """
    Parse one FSH file.
    Returns (class_id, description, elements) or None if not a Logical model.
    """
    text = path.read_text(encoding="utf-8")
    if "Logical:" not in text:
        return None

    id_m = re.search(r"^Id:\s+(\S+)", text, re.MULTILINE)
    class_id = id_m.group(1) if id_m else path.stem

    desc_m = _DESC.search(text)
    description = desc_m.group(1).strip() if desc_m else ""

    elements = []
    bindings: dict = {}

    for line in text.splitlines():
        if not line.startswith("*"):
            continue
        # Skip metadata (^) and root-element (. ^) annotations
        if re.match(r"^\*\s+[.^]", line):
            continue

        bm = _BIND.match(line)
        if bm:
            bindings[bm.group(1)] = _nos_name(bm.group(2))
            continue

        em = _ELEM.match(line)
        if em:
            elements.append({
                "name": em.group(1),
                "cardinality": f"[{em.group(2)}]",
                "type": _type_name(em.group(3)),
                "description": em.group(5),   # definition (second quoted string)
                "nomenclature": "",
            })

    # Attach ValueSet bindings to elements
    for elem in elements:
        nom = bindings.get(elem["name"], "")
        if not nom:
            nom = bindings.get(elem["name"].split(".")[-1], "")
        elem["nomenclature"] = nom

    return class_id, description, elements


def build_model_map() -> dict:
    """Scan all FSH files and return {class_id: {description, elements}}."""
    model_map: dict = {}
    for fsh_file in FSH_DIR.rglob("*.fsh"):
        result = parse_fsh(fsh_file)
        if result:
            class_id, description, elements = result
            model_map[class_id] = {"description": description, "elements": elements}
    return model_map


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def _add_row(ws, values, font=None, fill=None):
    ws.append(values)
    for cell in ws[ws.max_row]:
        cell.alignment = WRAP_TOP
        if font:
            cell.font = font
        if fill:
            cell.fill = fill


def generate(sections: dict, model_map: dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MOS"

    # 1. Column headers
    _add_row(ws, HEADERS, H_FONT, H_FILL)

    # 2. Partie overview rows (one per section)
    for partie in sections:
        _add_row(ws, [partie] + [""] * 14, P_FONT, P_FILL)

    # 3. Column headers repeated
    _add_row(ws, HEADERS, H_FONT, H_FILL)

    # 4. Class header rows
    for partie, classes in sections.items():
        for cid in classes:
            info = model_map.get(cid)
            if info is None:
                continue
            _add_row(
                ws,
                [partie, cid, None, None, None, info["description"]] + [""] * 9,
                C_FONT, C_FILL,
            )

    # 5. Column headers repeated
    _add_row(ws, HEADERS, H_FONT, H_FILL)

    # 6. Attribute rows
    n = 0
    for partie, classes in sections.items():
        for cid in classes:
            info = model_map.get(cid)
            if info is None:
                continue
            for e in info["elements"]:
                _add_row(
                    ws,
                    [
                        partie, cid,
                        e["name"], e["cardinality"], e["type"],
                        e["description"], e["nomenclature"],
                        "", "", "", "", "", "", "", "",
                    ],
                    fill=ODD_FILL if n % 2 == 0 else None,
                )
                n += 1

    # 7. Column widths & freeze panes
    for i, w in enumerate(
        [22, 30, 35, 13, 22, 100, 45, 35, 12, 12, 10, 12, 10, 14, 10], 1
    ):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    total_classes = sum(len(v) for v in sections.values())
    print(f"Saved {OUTPUT_FILE}  ({total_classes} classes, {n} attributes)")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    with open(SECTIONS_FILE, encoding="utf-8") as f:
        sections = json.load(f)
    model_map = build_model_map()
    generate(sections, model_map)
